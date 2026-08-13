#!/usr/bin/env python3
"""Normalize user-exported MoneyDJ NAV files without inventing missing data."""

from __future__ import annotations

import argparse
import csv
from datetime import date, datetime
from pathlib import Path

DATE_NAMES = {"日期", "淨值日期", "date", "nav date"}
NAV_NAMES = {"淨值", "基金淨值", "nav", "net asset value"}


def normalized_name(value: object) -> str:
    return str(value or "").strip().lower().replace("\ufeff", "")


def parse_date(value: object) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y.%m.%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    raise ValueError(f"Unsupported date: {text}")


def find_columns(headers: list[object]) -> tuple[int, int]:
    names = [normalized_name(item) for item in headers]
    date_index = next((i for i, name in enumerate(names) if name in DATE_NAMES), None)
    nav_index = next((i for i, name in enumerate(names) if name in NAV_NAMES), None)
    if date_index is None or nav_index is None:
        raise ValueError("找不到日期／淨值欄位")
    return date_index, nav_index


def read_rows(path: Path) -> list[list[object]]:
    if path.suffix.lower() == ".csv":
        for encoding in ("utf-8-sig", "cp950", "big5"):
            try:
                with path.open(encoding=encoding, newline="") as handle:
                    return list(csv.reader(handle))
            except UnicodeDecodeError:
                continue
        raise ValueError("無法辨識CSV編碼")
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("Excel匯入需要 openpyxl") from exc
        sheet = load_workbook(path, read_only=True, data_only=True).active
        return [list(row) for row in sheet.iter_rows(values_only=True)]
    raise ValueError("僅支援CSV、XLSX或XLSM")


def normalize_nav(path: Path) -> list[tuple[date, float]]:
    rows = read_rows(path)
    header_row = next((i for i, row in enumerate(rows) if any(normalized_name(x) in DATE_NAMES for x in row)), None)
    if header_row is None:
        raise ValueError("找不到表頭")
    date_index, nav_index = find_columns(rows[header_row])
    values: dict[date, float] = {}
    for row in rows[header_row + 1:]:
        if max(date_index, nav_index) >= len(row) or row[date_index] in (None, "") or row[nav_index] in (None, ""):
            continue
        try:
            values[parse_date(row[date_index])] = float(str(row[nav_index]).replace(",", ""))
        except (ValueError, TypeError):
            continue
    if len(values) < 2:
        raise ValueError("有效淨值資料不足2筆")
    return sorted(values.items())


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("input", type=Path)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()
    nav = normalize_nav(args.input)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    with args.output.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["date", "nav"])
        writer.writerows((day.isoformat(), value) for day, value in nav)
    print(f"{len(nav)} rows -> {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
