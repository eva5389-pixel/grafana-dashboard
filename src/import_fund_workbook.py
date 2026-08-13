#!/usr/bin/env python3
"""Import verified fund snapshot/performance fields from the supplied XLSM."""

from __future__ import annotations

import argparse
import csv
from datetime import date, datetime
from pathlib import Path


def usable_number(value):
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    return None


def category_for(region: object, target: object, name: str) -> str | None:
    text = " ".join(str(item or "") for item in (region, target, name))
    sector_rules = (
        ("quantum", ("量子",)),
        ("semiconductor", ("半導體",)),
        ("optical", ("光通訊", "光纖",)),
        ("healthcare", ("醫療", "健康", "生物科技", "生技")),
        ("finance", ("金融", "銀行", "保險")),
    )
    for category, words in sector_rules:
        if any(word in text for word in words):
            return category
    for category, words in (
        ("taiwan", ("台灣", "臺灣")),
        ("japan", ("日本",)),
        ("korea", ("韓國", "南韓")),
        ("hong_kong", ("香港",)),
        ("china", ("中國", "大中華")),
        ("usa", ("美國",)),
    ):
        if any(word in str(region or "") for word in words):
            return category
    return None


def overseas_rows(workbook):
    nav = workbook["淨值"]
    perf = workbook["報酬率"]
    perf_by_name = {}
    for row in perf.iter_rows(min_row=3, values_only=True):
        name = str(row[2] or "").strip()
        if name:
            perf_by_name[name] = row
    for row in nav.iter_rows(min_row=3, values_only=True):
        name = str(row[2] or "").strip()
        if not name:
            continue
        p = perf_by_name.get(name)
        if not p:
            continue
        category = category_for(row[3], row[4], name)
        if not category:
            continue
        yield {
            "category": category, "fund_code": "", "name": name,
            "as_of_date": row[1], "nav": row[5], "currency": row[6],
            "region": row[3], "target": row[4], "sharpe": row[8],
            "momentum_6m": p[8], "return_1y": p[9],
            "source_sheet": "淨值+報酬率",
        }


def domestic_rows(workbook):
    nav = workbook["國內淨值"]
    perf = workbook["國內報酬率"]
    perf_by_name = {}
    for row in perf.iter_rows(min_row=4, values_only=True):
        name = str(row[2] or "").strip()
        if name:
            perf_by_name[name] = row
    for row in nav.iter_rows(min_row=3, values_only=True):
        name = str(row[2] or "").strip()
        p = perf_by_name.get(name)
        if not name or not p:
            continue
        yield {
            "category": category_for("台灣", row[3], name) or "taiwan",
            "fund_code": "", "name": name, "as_of_date": row[1],
            "nav": row[4], "currency": row[5], "region": "台灣",
            "target": row[3], "sharpe": row[9],
            "momentum_6m": p[10], "return_1y": p[11],
            "source_sheet": "國內淨值+國內報酬率",
        }


def iso_date(value) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value or "")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("input", type=Path)
    parser.add_argument("--output", type=Path, default=Path("data/workbook_funds.csv"))
    args = parser.parse_args()
    from openpyxl import load_workbook
    workbook = load_workbook(args.input, read_only=True, data_only=True, keep_vba=True)
    rows = list(overseas_rows(workbook)) + list(domestic_rows(workbook))
    cleaned = []
    seen = set()
    for row in rows:
        key = (row["category"], row["name"])
        if key in seen:
            continue
        seen.add(key)
        for field in ("nav", "sharpe", "momentum_6m", "return_1y"):
            row[field] = usable_number(row[field])
        if row["nav"] is None or row["return_1y"] is None:
            continue
        row["as_of_date"] = iso_date(row["as_of_date"])
        cleaned.append(row)
    fields = ["category", "fund_code", "name", "as_of_date", "nav", "currency",
              "region", "target", "return_1y", "momentum_6m", "sharpe", "source_sheet"]
    args.output.parent.mkdir(parents=True, exist_ok=True)
    with args.output.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(cleaned)
    print(f"{len(cleaned)} verified snapshot rows -> {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
